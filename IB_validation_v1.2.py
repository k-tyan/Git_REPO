#!/usr/bin/env python3
'''
v1.1:
    - that version creates IB list and output file (tab "Detailed Inventory") containing duplicate SNs, no SN items excluded
    - creates tab "EoS Report" and lists SNs with EoS Flag set to "Y"
    - creates "SN to PN double mapping" tab that lists lines that have different Material Ids (PNs) associated with same SN
'''
    
import openpyxl
import sys
    
def main():
    
    try:
        in_file = sys.argv[1]
        out_file = sys.argv[2]
    except IndexError:    
        print('\nRun the script with input and output files:\n\t\t\tIB_validation.py <SBR report>.xlsx <output file>.xlsx\n\n')
    else:    
        workbook = in_file
        sheetName = 'Detailed Inventory'
        sheet = openpyxl.load_workbook(workbook)[sheetName]
        wb = sheet['A:AU']
        
        # exceptions_no_SN list lists PNs that have no SN such as power modules, fans, sw licenses etc
        exceptions_no_SN = ['ACX4000BASE-DC', 'CHAS-ACX4000-S', 'JUNOS-WW', 'JUNOS-WW-BB', 'JUNOS-64', 'JUNOS-LTD-64',
                      'JUNOS-WW-64', 'FANTRAY-MX80-BB', 'FFILTER-MX960-HC-BB', 'CBL-JX-PWR-EU', 'FFILTER-MX960-BB',
                      'FFANTRAY-MX480-HC-BB', 'CBL-M-PWR-RA-EU', 'CBL-PWR-10AC-STR-EU', 'S-MX80-Q', 'S-MPC-3D-VQ', 'MX2K-CBL-BTM-BB',
                      'JS-IPv6', 'S-MPC-3D-VQ-ADV-R', 'S-MPC-3D-PQ-ADV-R', 'S-SSM-FP', 'S-MX80-ADV-R', 'CBL-PWR-C15M-HITEMP-EU', 'CBL-M-PWR-RA-US',
                      'S-ACCT-JFLOW-IN', 'S-MPC3E-3D-ADV-R', 'S-SA-64K', 'S-SA-FP', 'S-MPC-3D-16XGE-ADV-R', 'S-ACCT-JFLOW-IN-5G',
                      'WLA-ANTPROT-OUT', 'PWR-T-BUS-BAR-S', 'MX2K-DCCBLMGR-BB', 'MX2K-EMI-BTM-BB',
                      'FAN-REAR-TX-T640-BB', 'FANTRAY-M10i-S', 'S-MPC4E-3D-ADV-R', 'S-MPC4E-3D-ADV-IR', 'MCG-PPB-S', 'S-MX80-SSM-FP',
                      'S-NAT', 'S-SA-4K', 'S-SA-16K', 'S-MX104-UPG-4X10GE', 'S-SFW', 'S-MX104-ADV-R2', 'S-MX104-Q', 'FAN-REAR-TXP-LCC-BB']
        
        # exceptions_sku dictionary lists PNs that are just another names for respective base PNs such as Premium bundles that include some line cards and modules that base PN does not
        exceptions_sku = {
            'NS-ISG-2000-SK1': 'NS-ISG-2000',
            'MX240BASE-DC': 'MX240-PREMIUM2-DC',
            'MX240BASE3-DC': 'MX240-PREMIUM3-DC',
            'MX240BASE-AC-HIGH': 'MX240-PREMIUM2-AC-HIGH',
            'MX480BASE-DC': 'MX480-PREMIUM2-DC',
            'MX480BASE3-DC': 'MX480-PREMIUM3-DC',
            'MX480BASE-AC': 'MX480-PREMIUM2-AC',
            'MX480BASE3-AC': 'MX480-PREMIUM3-AC',
            'MX960BASE3-DC': 'MX960-PREMIUM3-DC',
            'MX960BASE3-AC': 'MX960-PREMIUM3-AC'
        }
        
        exceptions_material_status = ['Scrapped', 'Inactive', 'None']
        
        headers = {'Serial Number': 0,
                   'Material Id': 0,
                   'Material Status': 0,
                   'EOS Flag': 0}
        
        def get_header_indx(ib_header_list, headers):
            for k in headers.keys():
                i = 0
                for cell in ib_header_list:
                    if cell == k:
                        headers[k] = i
                        break
                    else: i += 1
                    
            return headers
        
        # Finds line index in IB by SN and PN
        def find_line_in_ib(ib, headers, sn, pn):
            i = 0
            sn_index = get_header_indx(ib[0],headers)['Serial Number']
            pn_index = get_header_indx(ib[0],headers)['Material Id']
            while i < len(ib):
                if ib[i][sn_index] == sn and ib[i][pn_index] == pn:
                    break
                elif i == (len(ib) - 1):
                    i = 0
                    break
                else: i += 1
            return i
        
        def IB_init(sheet):
            ib = []
            
            flag = True
            for row in sheet.rows:
                sn_line = []
                if flag:
                    sn_line.append('DUP_FLAG')
                    flag = False
                else:
                    sn_line.append('')
                i = 0
                while i < len(row):
                    if row[i].value == None or row[i].value == '': sn_line.append('None')
                    else: sn_line.append(row[i].value)
                    i += 1
                    
                ib.append(sn_line)
            
            return ib
        
        # Returns an IB list with indication of duplicated SNs
        def mark_duplicates(ib, headers, exceptions_no_SN):
            sn_dict = {} #sn_dict contains unique combinations of SNs and PNs
            duplicates = {}
            sn_pn_confusion = []
            material_status_issues = []
            
            sn_index = get_header_indx(ib[0],headers)['Serial Number']
            pn_index = get_header_indx(ib[0],headers)['Material Id']
            item_status_index = get_header_indx(ib[0],headers)['Material Status']
            
            i = 0
            while i < len(ib):

                sn = ib[i][sn_index]
                pn = ib[i][pn_index]
                item_status = ib[i][item_status_index]

                if item_status not in exceptions_material_status:
                    if pn not in exceptions_no_SN:
                        if (sn, pn) not in sn_dict.items():
                            if sn not in sn_dict.keys(): sn_dict[sn] = pn
                            else:
                                n = find_line_in_ib(ib, headers, sn, sn_dict[sn])
                                sn_pn_confusion.append(ib[n])
                                sn_pn_confusion.append(ib[i])
                        elif (sn, pn) not in duplicates.items():
                            ib[i][0] = 'DUPLICATE'
                            n = find_line_in_ib(ib, headers, sn, pn)
                            ib[n][0] = 'duplicated'
                            duplicates[sn] = pn
                        else:
                            ib[i][0] = 'DUPLICATE'
                else: material_status_issues.append(ib[i])
                
                i += 1
            
            unique_sn_pn_pairs = len(sn_dict)    
            print('\nTotal number of unique SN-PN pairs is: {}\n'.format(unique_sn_pn_pairs))
            
            return ib, sn_pn_confusion, material_status_issues

        def sn_pn_confusion_cleanup_all(sn_pn_confusion, headers, exceptions_sku):
            l = list(sn_pn_confusion)
            for line in sn_pn_confusion:
                sn = line[headers['Serial Number']]
                pn = line[headers['Material Id']]
                if pn in exceptions_sku.keys():
                    cur = find_line_in_ib(l, headers, sn, pn)
                    nxt = find_line_in_ib(l, headers, sn, exceptions_sku[pn])
                    if nxt:
                        if nxt > cur:
                            print('nxt > cur which is expected')
                            del(l[nxt])
                            del(l[cur])
                        else:
                            #print('cur > nxt')
                            del(l[cur])
                            del(l[nxt])
                    #else:
                    #    print(sn)
                    #    print(pn)
            
            return l
            
                
        #===============================================MAIN================================================
        
        IB = IB_init(sheet)
        del(sheet)
        headers = get_header_indx(IB[0],headers)
        IB, SN_PN_confusion, Material_Status_Issues = mark_duplicates(IB, headers, exceptions_no_SN)
        
        # ==============================================OUTPUT==============================================    
        # import Workbook
        from openpyxl import Workbook
        # create Workbook object
        wb = Workbook()
        # set file path
        filepath = out_file
        # activate <file>.xlsx
        sheet = wb.active
        
        # Create, intialize and populate data into the new sheet "Duplicated SNs"
        sheet.title = 'Duplicated SNs'
        sheet.append(IB[0])
        sheet.auto_filter.ref = "A:BE"
        sheet.freeze_panes = 'A2'
        IB_dup_list = []
        for line in IB:
            if line[0] == 'duplicated' or line[0] == 'DUPLICATE':
                IB_dup_list.append(line)
        
        for line in IB_dup_list:
            sheet.append(line)
        
        # Create, intialize and populate data into the new sheet "SN to PN double mapping"
        wb.create_sheet(title = 'SN to PN double mapping')
        sheet = wb['SN to PN double mapping']
        sheet.append(IB[0])
        sheet.auto_filter.ref = "A:BE"
        sheet.freeze_panes = 'A2'
        SN_PN_confusion = sn_pn_confusion_cleanup_all(SN_PN_confusion, headers, exceptions_sku)
        for line in SN_PN_confusion:
            sheet.append(line)
       
        # Create, intialize and populate data into the new sheet "Material Status Issues"
        wb.create_sheet(title = 'Material Status Issues')
        sheet = wb['Material Status Issues']
        sheet.append(IB[0])
        sheet.auto_filter.ref = "A:BE"
        sheet.freeze_panes = 'A2'
        for line in Material_Status_Issues:
            sheet.append(line)

        # Create, intialize and populate data into the new sheet "EoS Report"
        wb.create_sheet(title = 'EoS Report')
        sheet = wb['EoS Report']
        sheet.append(IB[0])
        sheet.auto_filter.ref = "A:BE"
        sheet.freeze_panes = 'A2'
        eos_index = get_header_indx(IB[0],headers)['EOS Flag']
        for line in IB:
            if line[eos_index] == 'Y':
                sheet.append(line)
        
        # save file
        wb.save(filepath)

if __name__ == '__main__': main()

#=======================END OF SCRIPT====================================
